import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook

header=['rollno','register_sem','subno','sub_type']
def output_by_subject():
    dir_name =  "output_by_subject"
    if not os.path.exists(dir_name):
        os.mkdir(dir_name)
    data = open("regtable_old.csv", "r")
    data_csv=csv.reader(data)
    for record in data_csv:
        rollno,register_sem,schedule_sem,subno,grade1,date_of_entry1,grade2,date_of_entry2,sub_type = record
        refined_list = [rollno,register_sem,subno,sub_type]
        sub_no='{}.xlsx'.format(subno)
        file_path='./output_by_subject/'+sub_no
        if(os.path.isfile(file_path)): 
            workbook=load_workbook(r'output_by_subject\\{}.xlsx'.format(subno))
            sheet=workbook.active
            sheet.append(refined_list)
            workbook.save(r'output_by_subject\\{}.xlsx'.format(subno))                  
        else: 
            workbook=Workbook()
            sheet=workbook.active
            sheet.append(header)
            sheet.append(refined_list)
            workbook.save(f'output_by_subject\\{subno}.xlsx')
    return
 
def output_individual_roll():
    dir_name = "output_individual_roll"
    if not os.path.exists(dir_name):
        os.mkdir(dir_name)
    data = open("regtable_old.csv", "r")
    data_csv=csv.reader(data)
    data_list = [list(record) for record in data_csv][1:]  
    for record in data_list:
        rollno,register_sem,schedule_sem,subno,grade1,date_of_entry1,grade2,date_of_entry2,sub_type = record
        refined_list = [rollno,register_sem,subno,sub_type]
        roll_no='{}.xlsx'.format(rollno)
        file_path='./output_individual_roll/'+roll_no
        if(os.path.isfile(file_path)):
            workbook=load_workbook(r'output_individual_roll\\{}.xlsx'.format(rollno))
            sheet=workbook.active
            sheet.append(refined_list)
            workbook.save(r'output_individual_roll\\{}.xlsx'.format(rollno))                  
        else: 
            workbook=Workbook()
            sheet=workbook.active
            sheet.append(header)
            sheet.append(refined_list)
            workbook.save(f'output_individual_roll\\{rollno}.xlsx')
    return
 
output_by_subject()
output_individual_roll()