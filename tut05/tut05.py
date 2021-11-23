# RollNo : 1901EE34 . The code will be finished in a minute.So please wait for 1 min....
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import csv

def create_workbook(record): #creating workbook
    wb=Workbook()
    ws=wb.active
    ws.title = "Overall"
    wb.save(f'output\\{record[0]}.xlsx')
    return wb

def get_credits_spi(wb,credit_map): #getting spi and credit details
    Spi,Credits= [],[]
    for sheet in wb.sheetnames[1:]:
        ws = wb[sheet]
        credits = [int(cell.value) for cell in ws["E"][1:]]
        spi = [credit_map[cell.value.strip().strip("*")] for cell in ws["G"][1:]]
        Credits.append(sum(credits))
        Spi.append(round(sum([spi[i]*credits[i] for i in range(len(spi))])/sum(credits),2))
    return Spi,Credits

def get_results(Spi,ws,Credits,wb):        
    prefix_credits = [sum(Credits[:i+1]) for i in range(len(Credits))]
    CPI_num=[Spi[i]*Credits[i] for i in range(len(Credits))] 
    #appending all details
    ws.append(["Semester No."]+[x for x in range(1,len(wb.sheetnames))])
    ws.append(["Semester wise Credit Taken"]+Credits)
    ws.append(["SPI"]+Spi) 
    ws.append(["Total Credits Taken"]+prefix_credits)
    ws.append(["CPI"]+[round(sum(CPI_num[:i+1])/prefix_credits[i],2) for i in range(len(Spi))]) 
    return

def calculate_overallpage(wb,record,credit_map):
    name,Roll = record[1],record[0]
    ws =wb.active
    ws.column_dimensions["A"].width = 30
    ws.append(["Roll No.",Roll])
    ws.append(["Name of Student",name])
    ws.append(["Discipline",Roll[4:6]])
    Spi,Credits= get_credits_spi(wb,credit_map)
    get_results(Spi,ws,Credits,wb)
    return

def generate_marksheet():
    #checking if the directory exists
    print("Please wait for 1 min")
    dir_name =  "output"
    if not os.path.exists(dir_name):
        os.mkdir(dir_name)

    # loading all input csv modules
    course_data = open("subjects_master.csv", "r")
    course_csv=csv.reader(course_data)
    course_list =  [list(record) for record in course_csv][1:] 

    grades_data = open("grades.csv", "r")
    grades_csv = csv.reader(grades_data)
    grades_list = [list(record) for record in grades_csv][1:] 

    names_data = open("names-roll.csv","r")
    names_csv = csv.reader(names_data)
    names_list = [list(record) for record in names_csv]

    # storing grade-pointer values
    credit_map = {'AA':10,'AB':9,'BB':8,'BC':7,'CC':6,'CD':5,'DD':4,'F':0,'I':0,
                'AA*':10,'AB*':9,'BB*':8,'BC*':7,'CC*':6,'CD*':5,'DD*':4,'F*':0,'I*':0}
    
    # using map to store course details

    course_dict = {}
    max_length = 0

    for record in course_list:
        subno = record[0]
        if subno not in course_dict:
            course_dict[subno] = record
            max_length = max(max_length,len(record[1]))  
    
    #iterating through grades.csv
    c=1
    wb =0
    for i in range(len(grades_list)):
        record = grades_list[i]
        if i+1 != len(grades_list):         #checking if it is the last roww
            record1 = grades_list[i+1]

        Roll,Sem_no, = record[0],record[1]

        file_path='./output/'+'{}.xlsx'.format(Roll)

        if not os.path.isfile(file_path):   #checking if the {rollNo}.xlsx exists in output folder
            wb = create_workbook(record)    #function call for creating a workbook
        elif wb==0 :
            print("You are attempting to run the code second time which appends the same so plz delete the contents of output folder and run the code ")
            return 
        if f'Sem{int(Sem_no)}' not in wb.sheetnames: #checking if the sheet exists in {rollNo}.xlsx
            wb.create_sheet(f'Sem{Sem_no}',int(Sem_no))
            ws = wb[f"Sem{int(Sem_no)}"]
            ws.column_dimensions["C"].width = max_length      
        
        ws = wb[f"Sem{int(Sem_no)}"]

        if ws.max_row==1 :      #appending first row in sem sheets
            ws.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
        
        SubCode,Credit,Grade,Sub_Type = record[2:]
    
        subname,ltp,crd = course_dict[f"{SubCode}"][1:]   

        ws.append([ws.max_row,SubCode,subname,ltp,crd,Sub_Type,Grade]) # appending appropriate row in approriate sem sheet
        # print(f"Iterating {i+1} row")   

        if i+1==len(grades_list) or record[0]!=record1[0] :
            calculate_overallpage(wb,names_list[c],credit_map)  #filling overall page
            # print(f"Calculating {c} student's result")
            wb.save(r'output\\{}.xlsx'.format(Roll))         #saving after complete operations are done in rollNo.xlsx 
            c+=1;    
    print("Completed")
    return
generate_marksheet()